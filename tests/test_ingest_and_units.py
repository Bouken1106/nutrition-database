import requests
from openpyxl import Workbook

from src.db.connection import get_connection
from src.db.schema import create_schema
from src.ingest.estat import import_estat
from src.ingest.mext import import_mext
from src.ingest.open_food_facts import SEARCH_A_LICIOUS_URL, SEARCH_URL, search_products, upsert_off_product
from src.normalize.units import convert_unit, normalize_mass_to_g, parse_quantity_text


class StubResponse:
    def __init__(self, payload: dict[str, object], status_code: int = 200, url: str = "https://example.test") -> None:
        self._payload = payload
        self.status_code = status_code
        self.url = url

    def raise_for_status(self) -> None:
        if self.status_code >= 400:
            raise requests.HTTPError(f"{self.status_code} error for {self.url}", response=self)

    def json(self) -> dict[str, object]:
        return self._payload


class StubSession:
    def __init__(self, responses: dict[str, StubResponse]) -> None:
        self.responses = responses
        self.calls: list[str] = []

    def get(self, url: str, **_: object) -> StubResponse:
        self.calls.append(url)
        return self.responses[url]


def test_unit_normalization():
    assert normalize_mass_to_g(1, "kg") == 1000
    assert normalize_mass_to_g(500, "mg") == 0.5
    assert normalize_mass_to_g(500, "µg") == 0.0005
    assert parse_quantity_text("2 x 100 g") == (200.0, "g")
    assert convert_unit(2500, "mg", "g") == 2.5


def test_mext_import_from_excel(tmp_path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "MEXT"
    worksheet.append(["食品番号", "食品名", "エネルギー", "たんぱく質", "脂質", "炭水化物", "ビタミンC"])
    worksheet.append(["1001", "白米", 156, 2.5, 0.3, 37.1, 0])
    input_path = tmp_path / "mext.xlsx"
    workbook.save(input_path)

    db_path = tmp_path / "nutrition.db"
    with get_connection(db_path) as conn:
        create_schema(conn)
        imported = import_mext(conn, input_path)
        assert imported == 1
        food = conn.execute("SELECT * FROM foods WHERE source_type = 'mext'").fetchone()
        assert food["name"] == "白米"
        nutrients = conn.execute(
            "SELECT nutrient_id, amount_per_100g FROM food_nutrients WHERE food_id = ? ORDER BY nutrient_id",
            (food["food_id"],),
        ).fetchall()
        assert {row["nutrient_id"] for row in nutrients} >= {"energy_kcal", "protein_g", "carb_g"}


def test_mext_reimport_replaces_stale_nutrients(tmp_path):
    first_workbook = Workbook()
    first_sheet = first_workbook.active
    first_sheet.title = "MEXT"
    first_sheet.append(["食品番号", "食品名", "エネルギー", "たんぱく質"])
    first_sheet.append(["1001", "白米", 156, 2.5])
    first_path = tmp_path / "mext_first.xlsx"
    first_workbook.save(first_path)

    second_workbook = Workbook()
    second_sheet = second_workbook.active
    second_sheet.title = "MEXT"
    second_sheet.append(["食品番号", "食品名", "エネルギー"])
    second_sheet.append(["1001", "白米", 200])
    second_path = tmp_path / "mext_second.xlsx"
    second_workbook.save(second_path)

    db_path = tmp_path / "nutrition.db"
    with get_connection(db_path) as conn:
        create_schema(conn)
        import_mext(conn, first_path)
        import_mext(conn, second_path)
        nutrients = conn.execute(
            "SELECT nutrient_id, amount_per_100g FROM food_nutrients ORDER BY nutrient_id"
        ).fetchall()
        assert [(row["nutrient_id"], row["amount_per_100g"]) for row in nutrients] == [("energy_kcal", 200.0)]


def test_price_normalization_in_estat_import(tmp_path):
    input_path = tmp_path / "estat.csv"
    input_path.write_text(
        "name,price,quantity_value,quantity_unit,observed_at\n"
        "白米,250,1,kg,2026-04-01\n",
        encoding="utf-8",
    )
    db_path = tmp_path / "nutrition.db"
    with get_connection(db_path) as conn:
        create_schema(conn)
        imported = import_estat(conn, input_path)
        assert imported == 1
        row = conn.execute("SELECT * FROM food_prices").fetchone()
        assert row["price_yen"] == 250
        assert row["price_per_g"] == 0.25


def test_off_product_reimport_clears_missing_nutrients(tmp_path):
    db_path = tmp_path / "nutrition.db"
    with get_connection(db_path) as conn:
        create_schema(conn)
        upsert_off_product(
            conn,
            {
                "code": "4900000000000",
                "product_name": "オートミール",
                "nutriments": {"proteins_100g": 10},
            },
        )
        upsert_off_product(
            conn,
            {
                "code": "4900000000000",
                "product_name": "オートミール",
                "nutriments": {},
            },
        )
        nutrients = conn.execute("SELECT nutrient_id FROM food_nutrients").fetchall()
        assert nutrients == []


def test_off_search_prefers_search_a_licious_results():
    session = StubSession(
        {
            SEARCH_A_LICIOUS_URL: StubResponse(
                {
                    "hits": [
                        {
                            "code": "4900000000000",
                            "product_name": "オートミール",
                            "nutriments": {},
                        }
                    ]
                },
                url=SEARCH_A_LICIOUS_URL,
            )
        }
    )

    products = search_products("オートミール", session)

    assert [product["code"] for product in products] == ["4900000000000"]
    assert session.calls == [SEARCH_A_LICIOUS_URL]


def test_off_search_falls_back_to_legacy_when_search_a_licious_returns_empty():
    session = StubSession(
        {
            SEARCH_A_LICIOUS_URL: StubResponse({"hits": []}, url=SEARCH_A_LICIOUS_URL),
            SEARCH_URL: StubResponse(
                {
                    "products": [
                        {
                            "code": "4900000000001",
                            "product_name": "オートミール",
                            "nutriments": {},
                        }
                    ]
                },
                url=SEARCH_URL,
            ),
        }
    )

    products = search_products("オートミール", session)

    assert [product["code"] for product in products] == ["4900000000001"]
    assert session.calls == [SEARCH_A_LICIOUS_URL, SEARCH_URL]
