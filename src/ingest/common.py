from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


def read_excel_sheets(path: str | Path) -> list[tuple[str, list[tuple[object, ...]]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheets: list[tuple[str, list[tuple[object, ...]]]] = []
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        sheets.append((worksheet.title, rows))
    return sheets


def read_csv_rows(path: str | Path) -> list[dict[str, str]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader)


def write_csv(path: str | Path, fieldnames: Iterable[str], rows: Iterable[dict[str, object]]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fieldnames))
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

